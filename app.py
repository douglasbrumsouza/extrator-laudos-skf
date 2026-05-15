# ==============================================================================
# 1. CAIXA DE FERRAMENTAS (Importação de Bibliotecas)
# ==============================================================================
import streamlit as st
import streamlit.components.v1 as components
import zipfile, re, os, base64, json
from io import BytesIO
from datetime import datetime, timedelta

try:
    import pdfplumber
except ImportError:
    st.error("Erro: Instale o pdfplumber"); st.stop()

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    st.error("Erro: Instale o openpyxl"); st.stop()

try:
    import plotly.express as px
    import plotly.graph_objects as go
    import pandas as pd
except ImportError:
    st.error("Erro: Instale plotly e pandas"); st.stop()

try:
    import gspread
    from oauth2client.service_account import ServiceAccountCredentials
except ImportError:
    st.error("Erro: Instale gspread e oauth2client"); st.stop()

# ── CONFIGURAÇÕES DO GOOGLE SHEETS ─────────────────────────────────────────────
NOME_DA_PLANILHA = "Historico_Laudos_SKF"

def conectar_planilha():
    escopos = ['https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/drive']
    credenciais_dict = dict(st.secrets["gcp_service_account"])
    creds = ServiceAccountCredentials.from_json_keyfile_dict(credenciais_dict, escopos)
    cliente = gspread.authorize(creds)
    planilha = cliente.open(NOME_DA_PLANILHA).sheet1
    return planilha

def carregar_historico():
    try:
        planilha = conectar_planilha()
        return planilha.get_all_records()
    except Exception as e:
        st.error(f"Erro ao acessar o banco de dados no Google Sheets: {e}")
        return []

def salvar_novos_laudos(novos_laudos):
    planilha = conectar_planilha()
    dados_nuvem = planilha.get_all_records()
    
    ids_na_nuvem = [str(linha.get("ID de amostra", "")) for linha in dados_nuvem]
    
    lista_para_salvar = []
    for laudo in novos_laudos:
        if str(laudo.get("ID de amostra", "")) not in ids_na_nuvem:
            lista_para_salvar.append(laudo)
            
    if lista_para_salvar:
        cabecalhos = list(dados_nuvem[0].keys()) if dados_nuvem else list(lista_para_salvar[0].keys())
        linhas_formatadas = []
        for laudo in lista_para_salvar:
            linha = [laudo.get(col, "") for col in cabecalhos]
            linhas_formatadas.append(linha)
            
        planilha.append_rows(linhas_formatadas)
    
    return len(lista_para_salvar)

# ── LOGO DA EMPRESA (Texto em Base64) ─────────────────────────────────────────
LOGO_B64 = "data:image/webp;base64,UklGRtwNAABXRUJQVlA4TM8NAAAv/8F/ECo79v+rtqT00Ejd9/+U4O7u7u7u7u7u7u7u7u723rv37H32uff8sbPRCpHM4eLe3n0XUXt31oSVvwyLHF6KOzFhZe3dWQ+hKsTdMg+dzC2srDIsrJDsVdaLsEdgFeKuIYsZtL4IDz1qm0ATuku7zgCHsNJaTOBgtdptAm4pUhHuGrlF7v4Pccgu/vCUtiEQMwKXvHGXlEWkNQO01sWJHu4uKXFruHHL3F7j8ELcoXIcQocpWCILkmTTttbvvbdr9t7nPtu2bdu2bdu2bdu2bdu2/SRGkqRIkv5SnQL3/EsV1cPMzCABDAACTLJt27Zt27Zt27Zt27Zt27ZebQKU5TvpCNllkdZAyhmm7QHGGWI5iF0A5wx4BPAhHuAnAMcc4Hd8PKAPmbjQ2oMAJrwp/zu76exQys67Fv1P2HSiEINeYLIS4DpAADn8rQPcMIxVj2fSIrntqDLDo4QB4wE50zqhYetJmwOjZGW20kVeUQFtDug2EN8JHS9UFUNbpOcTzSZY1Ec4hjRqrXVCytYBdv6wRfkxT7OCdDLAe0JNgC8GMpth8kmNuTuoIaozOEUICnC69QRZVfYmgAeEpAbcZ6btKpNR9913cMO0wc4IUQGeMpOOibIpzN2BmbTmzghZdwaiRopUDI8sDD1CCPskw8glj4whIoAxHeAfoWx5yy9cEmCmDIjHhLaAvjAkVWRQuAFLCHWzFm65dCY5Ae4S8v4tHZ/81kph0r4Afwl9AQIAeqdYqHNR3uEB1hEKM9jaBd+OZBnGOyWgtwiJDbhpYFJYJB2aB+ANoXE8EwUtYRh1AX4RItdvSGtZgInW5REqA/w3jE7CAdKD0JmJAYKBMYgQOp9QYHQmlH68QHchpGbQRZiaU7EqFUxqC3Ke+gmtNwXSwkJceE+E2ADvPijAqQFuEGoD3GaBicgtZX+E3Aw2pfCqldAboCcnQPPWhF+tM8jNpcXSCcHbKZxHVkJxA13IoWNCcgYVTCscjCdY9tZTm2UYMwjNDekkk9KFk7k8PCtv++bkJUQHOJZixmIJ1UFS1QTmHiHWhWsA99bj37II2cvwq7J/Y9tbE/0Bkw6E7oZo40dCO/i2swTf2iWE/6hvr8Q4gBM+pfNMNkL5F/lSC879x4cWAT7jHMCnzsaKwsRlII0J6ZlRPy4mdmMdINvjSM8nWutYB/CH2SNKbGUQ2gM0ia0qvFt5LIkAX/EO4FtijNMT4u8zxt0wz0DHxGBwFvNeqZTK/TvM+13bSr2OUB8kBZR6PO5tValN4N6DlLo67gFcUYmt497vDmBI0xHyd5lPqvzYx0xLvQn7lvVA7DsLSFdgH4hlIPZgH8BOEOexL2ld2NcOwFvsA3iVC/sAPt4X+wB+lId95ZHrf9f/rv9d/7v+d/3v+t/1v+t/1/+u/13/u/53/e/63/X/ArLcjxKBUYytjGVeS8qjm1YbZMOY4YFXWYNqmBj99Noh+I9gYKEnqR1C9KdcaCTDSqMcV+E7/AGFvhTycet6DGM4vXPjWdWZwLqe7Y53dWf4/QvdjG0EECMeWpr7v9s8/t9kPam9o/SB4pPOxyR2jbSQc/hy6lfpAL8MkaqQ5r7I0UyqcepWXGP6dzbalmfyeNf9JJaAoTfZXXP/e9PyHlrnDJ8htU1pQWf8aGcYeHkT3jzQIva3fIjP6KbWtE2PfVErYFijlvgEN/T4lCmK7llaBydvIlsGWsjRTK7xmdq5hV2tNmAY0ZCSzCZzYMJn4jtHwvQxr4ePl6zFHNeyDp/p3VrZ1SxfHhwhL1njnpescc9rSOOe15DGPa8hjXteQ9pJm9nT3dTOL0xc7V5tzpvY9oG2lTl+PU3h6GTS+0YiTvvqKogZds8cmdrPkoE5f7/YWsYo6jvaVjL1KKWNxBxy+0zQMt2bKzJ5bAtadpb0HW0vw+yeI1Gnf3sTrExi98isEQ4q2FjSxHYMtM0Mon5MmEqcMm8iWwca97xsGve8bBr3vIluG2g0eKIAk9k/kVFSszqQuYNYM3++C/wGXi8mwJhmNyTklaCtOcLBJbsY48yGWJUEft2IZbTPba6/L30u4JNPxqiqBm3NOX279CKjwy56m9M15x8XoQZaKxL4UW9zuKZxccFplq8OHZF0wjVoa/5lgFUjZBf0rqOL85chtEpREEDUoxS2YniWXBZJJ5xZmzv7T6cRDysNsErkXXx7l8uzcLIPou4mtCxG0B4ks1EAaS1rWyCcWZs7w/ubnqZxkNUlE4gHTt53tLmZchKSJX1Hm5spCyFZ0ne0uZmyEJIlfUeb234WQrKuRlNRbe6MHmx6kspGtnW4imb74aR5PrzZkh2mpNNqc9vPQrbVtejGDO5uNO9Zvzl0L7HFARnPyo5AaVFt7owfbk5Okpzi8ZkVhtg6pfmfJKgoQ1LvFSdtUvtG2tyt5CRZLlSAwTVP+DeWuQ0Bxre6F1TU2ctMLjmV1ZAoB69Em5spJ0kzPNmDE37Jfj22fQEmsL4XVOjdjmxkacjtUhbfRr/LBpLMIMMwkJqRJUz9wkII0/tW1EcSLWtA1SIjH115GSc9oY29wTZNdOTzNqZxYaGDLRlO98aqZDOIKE+TFoqGJUJEZZXJyYqOg57K6ZlJVKKFdDQsFTqi46VLNMmzkq4tTSq9yuJyvvqU3yMhXVsa9una0rBP15aGfbq2NAdoVYGdjoadn2MGeDoadnwWvpsAT9eW5vRQd+JbIpcJ7HQ07PRIMNTXwr6j25uOhqU3x2+XQTaIhcnnW33CaSDqZ6mAzeloWHajGFMhv6d0YuY47MrudDQsuR369yABIkHWY21PR8Nyu5d//a8YEmAPTpuOhu2Fls4tc1mOm46G7YUGXjc21vmt6/Mcx+LW4FukyiLnTdeWZi9OY2Cga0vDPl1bGsL0MLlNLjoatlBuPkPtlHGu8oxhRkMy+uaHtUzaHfhM+9qqW3GMICQ0sDqx7eTgecLF2EKe3S6hGS3d2X08LWtU4ys5uB588C1TnzGla9HUMLvlCs382U5znsWLw7iWd4bRNXerICNck+Zc/XtsoKxsWtBvZhNKwJP1OKXNhEX9WQt5zF5kcAQXN9bcZ/v+tCv53UOLevf0IVWRVHRDJlxHC1pbcDGFYzN+ut9lAvIrU5h3UaQiqTy8LP+menYhSp2RoOLPAtxLfs0Ks2GiUCGZ6NX714AoeldOXYNEFCqEef0oESAiChVCvL4X88WgUHa8+3ksFMqOdn0u6IuNQtmxrqU4KJQd6YqJi0LZca73uT0+UCg7yh3eF4pMaNMA417oE0UKIlyvs7l8o0hBfOtVZpcfFCmIbiX7Q5GC2Nbz9A6/KFIQ2RbjH0WScS2nCeQ15IQNt09BrF9L7YZmkNeQA5ZQpEjdiGlMcs9IagVMIa8h50vP9ddlWldW9QlbxVa11Eo0h7yGnC/pWuiwJlEkGc26m9BiFkWSsSyvaRRJRrJbmUeRZBxbLwfyklHsCjzIS8awz3AhL9mpmdLJmQ31LJ3Dv0OLUn2ID3klCHehkH8/EaXOhQcVD7KfreTxbxuiZCbeXgmiZSb/2xXlDhRUPnYDdjP7z6cGyf9QDWIkHJQbeRPa2BMr4aAmpK1PjBV9Nbggb4htUtUewk6vkoXMzHOvux6C+5jnNE7Y7XgWEjBPR2OY2TiEsNfbfF4yMzTQ2rHt5MiRYwmH4PrkYfXI5SXX/67/Xf+7/nf97/rf9b/rf9f/rv9d/7v+d/3v+t/1v+v/BXQ6FftSN4V99ReNfbkugX0Ab/+NfSAeM+Qy9p0f4AD2MZN9YKzBPjBWgzEN+z7JkO7YdyxDVMc+Q1Llftj3opWl4l55XYv+J6z6G+49UymALbgHsFEpgOG4B6ZDlDKgIu4ZpqWVajgV81KZPaIopc6IeVdXSimABZgHpjNjLBbzAK0UI4PnRgb4i3cAASuLoQBO4B1Dj6hYAfrg3bFiMzySol0qC0yC2NT5sa5ZFefjse5mcbGwT9yacG5TTESPS1WKc00oHxlUwLmd+tIJ9xXsrRh3xn66O4gvCtCuGMdEa+XzqVvFN4D3uX1TgIzFt3zKz0U/BdtaLeCPAnQ0toEYqPzuIq+ouXBtT13w7Uj+qVpxbc3KxE46QpaOaf9YjxmKiWqYll+ZC2IPnp1emZzOfyV6CpYV3aUw4pmlDohlb1KmM3cHbRbHbly1earLniTJhWGvqVjxBLQhhl1W8QVYil+1KM7M+aH/j11gXNw2L9Ul30qocWtPzHsnVvyZJ8V/h1mbSicKKRHviVipj1FirhWv8ilBU/6DVctXwqYYMA+nslYtjkp4DkY18XklcsJS8MkQyz+vxE7ZHzZdvWolekq+VExKzaesyKQ174tHm3qMsiYTBTUWXaJ6ZdWzA5zEoRUeT1k3sRYMqiVRWfqjl8CeVsGop6zOIuxYDHbgDoPjxSjrp4CkVat4k+tNVSspriAr1lRasZLmkQCuY8w/8iuZfh5EW40te1rzepRkcwP03BOmFA0wqoCS8MpA0h/gNZZcIt/+laQTH8PgGoY882a5lcRTXveg++LGpgDW7rRqJfsChmgDcKg8rEjd9c0aVvbYpfBn7LuA2HNffGg977EqVnbadsefPH9NeJB6dTCd2VzhyoZbhAgpVutzAG607vSV98yVg+mQjgsoO098Z/43AYx7UGtFrCs+fsHOWevx8WA8AbgEsL+JTx5rry/KrSwPAA=="

# ── Config ────────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Análise de Óleo SKF — Gerdau",
    page_icon="🛢️",
    layout="wide",
    initial_sidebar_state="collapsed"
)

# ── CSS global (O "Decorador" Visual do Dashboard) ─────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Barlow:wght@400;500;600;700&family=Barlow+Condensed:wght@600;700&display=swap');

/* Esconde as coisas feias padrão do Streamlit */
html, body, [class*="css"] { font-family: 'Barlow', sans-serif !important; }
#MainMenu, footer, header { display: none !important; }
[data-testid="stHeader"]    { display: none !important; }
[data-testid="stToolbar"]   { display: none !important; }
[data-testid="stDecoration"]{ display: none !important; }
[data-testid="stStatusWidget"] { display: none !important; }
section[data-testid="stSidebar"] { display: none !important; }

.block-container {
    padding: 2.5rem 1.5rem 1rem 1.5rem !important; 
    max-width: 100% !important;
    margin-top: 0 !important;
}

/* Fundo Corporate Clean */
.stApp { background: #F8FAFC !important; }

div[data-testid="stVerticalBlock"] { gap: 0.6rem !important; }

/* ── DASHBOARD: CABEÇALHO E BARRAS ── */
.db-header {
    background: #0055A5;
    border-radius: 4px; padding: 12px 18px; 
    margin-bottom: 12px !important; 
    display: flex; align-items: center; justify-content: space-between;
    box-shadow: 0 4px 6px -1px rgba(0,0,0,0.1);
}
.db-logo {
    width: 42px; height: 42px;
    background: transparent !important; 
    display: flex; align-items: center; justify-content: center;
    flex-shrink: 0;
}
.db-logo img {
    width: 100%; height: 100%; object-fit: contain;
}
.db-title {
    font-family: 'Barlow Condensed',sans-serif;
    font-size: 20px; font-weight: 700; color: white; letter-spacing: 0.8px;
}
.db-sub { font-size: 11px; color: rgba(255,255,255,0.85); margin-top: 1px; }
.db-badge { 
    font-size: 9px; font-weight: 600; color: #0055A5;
    background: #FFFFFF; padding: 2px 10px;
    border-radius: 2px;
}

.act-bar {
    background: #FFFFFF; border: 1px solid #E2E8F0; border-radius: 4px;
    padding: 0 16px; display: flex; align-items: center; gap: 10px;
    height: 42px; box-shadow: 0 1px 3px rgba(0,0,0,0.02);
}
.act-title {
    font-family: 'Barlow Condensed',sans-serif; font-size: 14px;
    font-weight: 700; color: #0F172A; flex: 1;
}
.bcnt { 
    font-size: 10px; font-weight: 700; padding: 2px 8px;
    border-radius: 2px; white-space: nowrap;
}

/* ── DASHBOARD: OS CARTÕES DE RESUMO (KPIs) ── */
.kpi { 
    border-radius: 4px; padding: 16px 10px; text-align: center;
    background: #FFFFFF !important; border: 1px solid #E2E8F0 !important; 
    position: relative; overflow: hidden; height: 100%; box-sizing: border-box; 
    box-shadow: 0 2px 5px rgba(0,0,0,0.03); 
}
.kpi::before { content:''; position:absolute; top:0; left:0; bottom:0; width:6px; height:100%; border-radius: 4px 0 0 4px; }
.kpi-tot::before { background:#0055A5; }
.kpi-nor::before { background:#10B981; }
.kpi-alt::before { background:#F59E0B; }
.kpi-alm::before { background:#EF4444; }

.kv { font-family:'Barlow Condensed',sans-serif; font-size:32px; font-weight:700; line-height:1; } 
.kpi-tot .kv { color:#0055A5; }
.kpi-nor .kv { color:#10B981; }
.kpi-alt .kv { color:#F59E0B; }
.kpi-alm .kv { color:#EF4444; }
.kl { font-size:10px; font-weight:600; text-transform:uppercase; letter-spacing:1px; margin-top:6px; color:#475569 !important; }
.ks { font-size:10px; margin-top:2px; color:#64748B !important; }

/* ── CAIXAS DOS GRÁFICOS E TABELA ── */
.sc { background:#FFFFFF; border:1px solid #E2E8F0; border-radius:4px; padding:12px 14px; box-shadow: 0 1px 3px rgba(0,0,0,0.02); }

.sc-top { 
    background:#FFFFFF; border:1px solid #E2E8F0; border-bottom:none;
    border-radius:4px 4px 0 0; padding:12px 14px 20px 14px;
    position: relative; z-index: 1;
}
div[data-testid="stPlotlyChart"] { 
    background: #FFFFFF; border: 1px solid #E2E8F0; border-top: none;
    border-radius: 0 0 4px 4px; padding: 0 10px 10px 10px;
    margin-top: -24px !important;
    position: relative; z-index: 10;
    box-shadow: 0 1px 3px rgba(0,0,0,0.02);
}

.st { 
    font-size:11px; font-weight:700; text-transform:uppercase; letter-spacing:1px;
    color:#0F172A; display:flex; align-items:center; gap:6px; margin-bottom:8px; 
}
.st::before { content:''; display:inline-block; width:3px; height:12px;
              border-radius:2px; background:#0055A5; } 

/* ── TABELA FINAL ── */
.tbl { width:100%; border-collapse:collapse; font-size:11px; }
.tbl thead tr { background:#F8FAFC; }
.tbl thead th { padding:8px 10px; text-align:left; font-size:9px; font-weight:700;
                text-transform:uppercase; letter-spacing:.7px; color:#475569;
                border-bottom:2px solid #E2E8F0; }
.tbl tbody tr { border-bottom:1px solid #E2E8F0; }
.tbl tbody tr:hover { background:#F1F5F9; } 
.tbl tbody td { padding:8px 10px; color:#334155; }
.td-eq { color:#0F172A !important; font-weight:600 !important; }
.td-set { font-family:'Barlow Condensed',sans-serif; font-size:13px; font-weight:700; color:#0055A5 !important; }
.td-par { color:#D97706 !important; font-weight:500; }
.td-obs { font-size:10px; color:#64748B !important; font-style:italic; }
.td-data { font-size:10px; color:#64748B !important; }
.td-red { color:#EF4444 !important; font-weight:700 !important; }
.td-yel { color:#F59E0B !important; font-weight:700 !important; }
.badge { display:inline-block; font-size:9px; font-weight:700; padding:2px 8px;
         border-radius:2px; text-transform:uppercase; letter-spacing:.4px; }
.ba { background:rgba(239,68,68,.1); color:#EF4444; border:1px solid rgba(239,68,68,.3); }
.bl { background:rgba(245,158,11,.1); color:#D97706; border:1px solid rgba(245,158,11,.3); }
.bn { background:rgba(16,185,129,.1); color:#10B981; border:1px solid rgba(16,185,129,.3); }

/* ── CAIXAS DE FILTROS COM DESTAQUE ── */
div[data-testid="stSelectbox"] label p {
    color:#475569 !important; font-size:13px !important; font-weight:700 !important;
    text-transform:uppercase; letter-spacing:0.5px;
}
div[data-testid="stSelectbox"] div[data-baseweb="select"] > div {
    background-color: #FFFFFF !important;
    border: 1px solid #CBD5E1 !important;
    border-radius: 6px !important;
    box-shadow: 0 2px 8px rgba(0,0,0,0.04) !important;
    transition: all 0.2s ease; 
}
div[data-testid="stSelectbox"] div[data-baseweb="select"] > div:hover {
    border-color: #0055A5 !important;
    box-shadow: 0 4px 12px rgba(0,85,165,0.1) !important;
}

/* Oculta as margens nos elementos nativos de upload para encaixar no card */
div[data-testid="stFileUploader"],
div[data-testid="stSpinner"],
div[data-testid="stAlert"],
div.stProgress {
    width: 100% !important;
    max-width: 540px !important;
    margin: 10px auto !important;
}
div[data-testid="stButton"] { display: flex; justify-content: center; }
div[data-testid="stButton"] > button { max-width: 540px; }
</style>
""", unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════════════
# EXTRAÇÃO (A "Fábrica" que lê o PDF)
# ══════════════════════════════════════════════════════════════════════════════
PARAMETROS = [
    ("Ferro","Ferro (ppm)"),("Crômio","Crômio (ppm)"),("Níquel","Níquel (ppm)"),
    ("Alumínio","Alumínio (ppm)"),("Chumbo","Chumbo (ppm)"),("Cobre","Cobre (ppm)"),
    ("Estanho","Estanho (ppm)"),("Titânio","Titânio (ppm)"),("Prata","Prata (ppm)"),
    ("Antimônio","Antimônio (ppm)"),("Cádmio","Cádmio (ppm)"),("Manganês","Manganês (ppm)"),
    ("Bismuto","Bismuto (ppm)"),("Arsênio","Arsênio (ppm)"),("Índio","Índio (ppm)"),
    ("Cobalto","Cobalto (ppm)"),("Zircônio","Zircônio (ppm)"),("Tungstênio","Tungstênio (ppm)"),
    ("Cério","Cério (ppm)"),("Silício","Silício (ppm)"),("Sódio","Sódio (ppm)"),
    ("Vanádio","Vanádio (ppm)"),("Potássio","Potássio (ppm)"),("Lítio","Lítio (ppm)"),
    ("água","Água (ppm)"),("Bolhas","Bolhas"),("Molibdênio","Molibdênio (ppm)"),
    ("Cálcio","Cálcio (ppm)"),("Magnésio","Magnésio (ppm)"),("Fósforo","Fósforo (ppm)"),
    ("Zinco","Zinco (ppm)"),("Bário","Bário (ppm)"),("Boro","Boro (ppm)"),
    ("TAN","TAN (mg KOH/g)"),("Oxidação","Oxidação (abs/0.1mm)"),
    ("Visc 40","Visc 40 (cSt)"),("Integridade de Fluído","Integridade de Fluído"),
]

def to_num(v):
    if v is None: return None
    s = str(v).strip().replace(",", ".")
    try: return float(s)
    except: return s if s else None

def detectar_status(pdf_bytes):
    with pdfplumber.open(BytesIO(pdf_bytes)) as pdf:
        full = "".join(p.extract_text() or "" for p in pdf.pages)
        if "Limites acionados" not in full: return "Normal"
        for img in pdf.pages[0].images:
            if img.get("name") == "Im2":
                return "Alarme" if img["height"] >= 50 else "Alerta"
    return "Alerta"

def split_ativo(raw):
    p = [x.strip() for x in raw.split("→")]
    s = p[0].strip().split("-")
    c = {f"Cod{i}": s[i-1] if len(s) >= i else "" for i in range(1, 6)}
    c["Equipamento Descrição 1"] = p[1] if len(p) > 1 else ""
    c["Equipamento Descrição 2"] = p[2] if len(p) > 2 else ""
    return c

def split_data(s):
    if not s: return None, None, None, None
    m = re.match(r"(\d{2})/(\d{2})/(\d{4}),?\s*(\d{2}:\d{2}:\d{2})?", str(s).strip())
    if not m: return None, None, None, None
    return int(m.group(1)), int(m.group(2)), int(m.group(3)), m.group(4)

def bloco(t, ini, fim):
    m = re.search(rf"{re.escape(ini)}\s*\n(.*?)(?={re.escape(fim)})", t, re.DOTALL)
    return m.group(1).strip() if m else ""

def extrair_laudo(nome, pdf_bytes):
    dados = {"Arquivo": nome}
    with pdfplumber.open(BytesIO(pdf_bytes)) as pdf:
        pags = [p.extract_text() or "" for p in pdf.pages]
    p1 = pags[0]; texto = "\n".join(pags)
    
    dados["Status"] = detectar_status(pdf_bytes)
    
    m = re.search(r"Instalação:\s*(.+?)(?:\s{2,}|\n)", p1)
    dados["Instalação"] = m.group(1).strip() if m else None
    
    m = re.search(r"Localização:\s*(.+)", p1)
    loc = m.group(1).strip() if m else ""
    if loc.endswith("-"):
        m2 = re.search(r"Localização:.*\n(.+?)(?:\n|$)", p1)
        if m2: loc = loc + " " + m2.group(1).strip()
    dados["Localização"] = loc
    
    m = re.search(r"Ativo:\s*(.+?)(?=Número de série:|\n\n)", p1, re.DOTALL)
    raw = re.sub(r"\s+", " ", m.group(1)).strip() if m else ""
    dados["Ativo (completo)"] = raw; dados.update(split_ativo(raw))
    
    m = re.search(r"Tipo de componente:\s*(.+?)(?=Operação de ativo:|\n)", p1)
    dados["Tipo de componente"] = m.group(1).strip() if m else None
    
    m = re.search(r"Óleo:\s*(.+?)(?=Operação de óleo:|\n)", p1)
    dados["Óleo"] = m.group(1).strip() if m else None
    
    m = re.search(r"data de coleta:\s*([\d/,: ]+)", p1)
    ds = m.group(1).strip() if m else None
    dados["Data de coleta"] = ds
    d, me, a, h = split_data(ds)
    dados["Dia coleta"] = d; dados["Mês coleta"] = me
    dados["Ano coleta"] = a; dados["Hora coleta"] = h
    
    dados["Observações"]              = bloco(p1, "Observações:", "Diagnósticos:")
    dados["Diagnósticos"]             = bloco(p1, "Diagnósticos:", "Ações:")
    dados["Ações"]                    = bloco(p1, "Ações:", "Recomendações adicionais:")
    dados["Recomendações adicionais"] = bloco(p1, "Recomendações adicionais:", "Teste realizado por:")
    
    m = re.search(r"Teste realizado por:\s*(.+?)(?=Lançado por:|\n)", p1)
    dados["Teste realizado por"] = m.group(1).strip() if m else None
    
    m = re.search(r"Lançado por:\s*(.+?)(?=_{3,}|\n)", p1)
    dados["Lançado por"] = m.group(1).strip() if m else None
    
    m = re.search(r"ID de amostra\s+([\w-]+)", p1)
    dados["ID de amostra"] = m.group(1).strip() if m else None
    
    m = re.search(r"Data de teste\s+(\d{2}/\d{2}/\d{4},\s*\d{2}:\d{2}:\d{2})", p1)
    dt = m.group(1).strip() if m else None
    dados["Data de teste"] = dt
    d, me, a, _ = split_data(dt)
    dados["Dia teste"] = d; dados["Mês teste"] = me; dados["Ano teste"] = a
    
    m = re.search(r"Data de lançamento\s+(\d{2}/\d{2}/\d{4},\s*\d{2}:\d{2}:\d{2})", p1)
    dados["Data de lançamento"] = m.group(1).strip() if m else None
    
    linhas = texto.split("\n"); pm = {}
    for linha in linhas:
        ls = linha.strip()
        for raw, _ in PARAMETROS:
            if raw not in pm and (ls.startswith(raw + " ") or ls == raw):
                parts = ls.split(); nw = len(raw.split())
                if len(parts) > nw:
                    v = parts[nw]
                    if re.match(r"^[\d\-]", v) or v in ("Normal", "Desconhecido"):
                        pm[raw] = v
                break
    for raw, col in PARAMETROS:
        dados[col] = to_num(pm.get(raw))
        
    m = re.search(r"Limites acionados \(amostra atual\)\n(.+?)(?=TruVu|Relatório)", texto, re.DOTALL)
    if m:
        pa = []
        for l in m.group(1).strip().split("\n"):
            pm2 = re.match(r"^\d+\s+(.+?)\s+(?:Absoluto|Desvio)", l.strip())
            if pm2: pa.append(pm2.group(1).strip())
        dados["Parâmetros em Alarme"] = "; ".join(pa)
    else:
        dados["Parâmetros em Alarme"] = ""
        
    return dados

def gerar_excel(lista):
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Laudos"
    colunas = list(lista[0].keys())
    cor_st = {"Normal": "C6EFCE", "Alerta": "FFEB9C", "Alarme": "FFC7CE"}
    hf = PatternFill("solid", start_color="0055A5") # Cor Azul Oficial SKF para o Excel
    hfont = Font(bold=True, color="FFFFFF", name="Arial", size=9)
    borda = Border(left=Side(style="thin"), right=Side(style="thin"),
                   top=Side(style="thin"), bottom=Side(style="thin"))
    for ci, col in enumerate(colunas, 1):
        c = ws.cell(row=1, column=ci, value=col)
        c.fill = hf; c.font = hfont; c.border = borda
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 36
    
    for ri, linha in enumerate(lista, 2):
        st_val = linha.get("Status", "Normal")
        fb = PatternFill("solid", start_color="F8FAFC" if ri % 2 == 0 else "FFFFFF")
        for ci, col in enumerate(colunas, 1):
            c = ws.cell(row=ri, column=ci, value=linha.get(col))
            c.font = Font(name="Arial", size=9); c.border = borda
            c.alignment = Alignment(horizontal="center", vertical="center")
            if col == "Status":
                c.fill = PatternFill("solid", start_color=cor_st.get(st_val, "FFFFFF"))
                c.font = Font(name="Arial", size=9, bold=True)
            else: c.fill = fb
            
    larg = {"Arquivo":35,"Status":10,"Localização":28,"Ativo (completo)":45,
            "Cod1":7,"Cod2":7,"Cod3":8,"Cod4":9,"Cod5":7,
            "Equipamento Descrição 1":26,"Equipamento Descrição 2":26,
            "Tipo de componente":28,"Óleo":24,"Data de coleta":20,
            "Dia coleta":8,"Mês coleta":8,"Ano coleta":8,"Hora coleta":10,
            "Observações":40,"Diagnósticos":50,"Ações":50,"Parâmetros em Alarme":30}
    for ci, col in enumerate(colunas, 1):
        ws.column_dimensions[get_column_letter(ci)].width = larg.get(col, 14)
        
    ws.freeze_panes = "A2"; ws.auto_filter.ref = ws.dimensions 
    buf = BytesIO(); wb.save(buf); buf.seek(0)
    return buf.read() 

# ══════════════════════════════════════════════════════════════════════════════
# TELA DE UPLOAD (A Tela de Recepção Protegida)
# ══════════════════════════════════════════════════════════════════════════════
def render_upload():
    # Usamos o 'components.html' aqui para o Streamlit não bloquear o design
    components.html(f"""
    <!DOCTYPE html>
    <html>
    <head>
    <meta charset="UTF-8">
    <link href="https://fonts.googleapis.com/css2?family=Barlow:wght@400;500;600;700&family=Barlow+Condensed:wght@600;700&display=swap" rel="stylesheet">
    <style>
    * {{ margin:0; padding:0; box-sizing:border-box; }}
    body {{
        font-family: 'Barlow', sans-serif;
        background: transparent;
        display: flex; flex-direction: column; align-items: center; justify-content: flex-end; padding: 20px 20px 0 20px;
    }}
    .card {{
        width: 100%; max-width: 540px;
        background: #FFFFFF; border-radius: 4px; overflow: hidden;
        box-shadow: 0 4px 20px rgba(0,0,0,0.08); border: 1px solid #E2E8F0;
    }}
    .head {{ background: #0055A5; padding: 26px 30px; }}
    .logo-row {{ display:flex; align-items:center; gap:14px; margin-bottom:12px; }}
    .logo {{
        width:48px; height:48px; background:transparent; 
        display:flex; align-items:center; justify-content:center;
    }}
    .title {{ font-family:'Barlow Condensed',sans-serif; font-size:22px; font-weight:700; color:white; }}
    .sub {{ font-size:11px; color:rgba(255,255,255,0.8); margin-top:2px; }}
    .desc {{ font-size:13px; color:rgba(255,255,255,0.9); line-height:1.6; }}
    .body {{ padding: 24px 30px; }}
    .steps {{
        background:#F8FAFC; border:1px solid #E2E8F0; border-radius:4px;
        padding:14px 18px; font-size:13px; color:#334155; line-height:2.2;
    }}
    .steps b {{ color:#0F172A; }}
    .foot {{ text-align:center; font-size:10px; color:#94A3B8; padding: 14px 0 20px 0; letter-spacing:.4px; }}
    </style>
    </head>
    <body>
    <div class="card">
      <div class="head">
        <div class="logo-row">
          <div class="logo">
            <img src="{LOGO_B64}" style="width:100%; height:100%; object-fit:contain;">
          </div>
          <div>
            <div class="title">Extrator de Laudos SKF</div>
            <div class="sub">Gerdau Charqueadas · Eng. de Manutenção</div>
          </div>
        </div>
        <div class="desc">
          Faça o upload do ZIP com os laudos em PDF.<br>
          Gera o <strong>Excel consolidado</strong> e o <strong>painel interativo</strong> automaticamente.
        </div>
      </div>
      <div class="body">
        <p style="font-size:12px; color:#475569; margin-bottom:10px;">
          O sistema está conectado à planilha <b>{NOME_DA_PLANILHA}</b>.
        </p>
        <div class="steps">
          📦 &nbsp;<b>1.</b> Receba o ZIP do laboratório SKF<br>
          ⬆️ &nbsp;<b>2.</b> Selecione o arquivo abaixo<br>
          ⚙️ &nbsp;<b>3.</b> Clique em Processar Laudos<br>
          📊 &nbsp;<b>4.</b> Veja o painel e baixe o Excel
        </div>
      </div>
      <div class="foot">Desenvolvido por Douglas Brum · SKF</div>
    </div>
    </body>
    </html>
    """, height=460)

    uploaded = st.file_uploader(
        "📦 Selecione o arquivo ZIP com os laudos",
        type=["zip"],
        label_visibility="collapsed"
    )
    
    if uploaded:
        st.success(f"✅ **{uploaded.name}** — {uploaded.size/1024/1024:.1f} MB") 
        
        if st.button("⚙️  Processar Laudos", type="primary", use_container_width=True):
            with st.spinner("Extraindo e enviando para o histórico..."): 
                zf = zipfile.ZipFile(BytesIO(uploaded.read())) 
                
                pdfs = [(n, zf.read(n)) for n in zf.namelist()
                        if n.lower().endswith(".pdf") and not n.startswith("__MACOSX")]
                        
            if not pdfs:
                st.error("Nenhum PDF encontrado no ZIP."); return 
                
            prog = st.progress(0) 
            lista_novos = []
            
            for i, (nome, pdf_bytes) in enumerate(sorted(pdfs), 1):
                prog.progress(i / len(pdfs),
                    text=f"Processando {i}/{len(pdfs)}: {os.path.basename(nome)[:45]}")
                try: lista_novos.append(extrair_laudo(os.path.basename(nome), pdf_bytes))
                except: pass
                
            prog.progress(1.0, text="Salvando no Banco de Dados...")
            
            qtd_salvos = salvar_novos_laudos(lista_novos)
            st.success(f"{qtd_salvos} laudos novos gravados no histórico com sucesso!")
            
            st.session_state["laudos"] = carregar_historico()
            st.session_state["processado"] = True
            st.session_state["excel"] = f"LAUDOS_COMPLETO_{datetime.now().strftime('%d-%m-%Y')}.xlsx"
            st.rerun() 
    else:
        st.markdown("""
        <p style="text-align:center;color:#64748B;font-size:12px;margin-top:4px;width:100%;max-width:540px;margin-left:auto;margin-right:auto;">
        Arraste o arquivo ZIP aqui ou clique para selecionar
        </p>""", unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════════════
# DASHBOARD
# ══════════════════════════════════════════════════════════════════════════════
def render_dashboard(lista):
    df = pd.DataFrame(lista) 
    
    MESES = {1:"Jan",2:"Fev",3:"Mar",4:"Abr",5:"Mai",6:"Jun",
             7:"Jul",8:"Ago",9:"Set",10:"Out",11:"Nov",12:"Dez"}
    MESES_F = {1:"Janeiro",2:"Fevereiro",3:"Março",4:"Abril",5:"Maio",6:"Junho",
               7:"Julho",8:"Agosto",9:"Setembro",10:"Outubro",11:"Novembro",12:"Dezembro"}

    total_g = len(df)
    norm_g  = len(df[df["Status"]=="Normal"])
    alt_g   = len(df[df["Status"]=="Alerta"])
    alm_g   = len(df[df["Status"]=="Alarme"])

    excel_bytes = gerar_excel(lista) 
    
    # ── BARRA SUPERIOR ──
    c1, c2, c3, c4 = st.columns([3.5, 1.2, 1.1, 1.2]) 
    with c1:
        st.markdown(f"""
        <div class="act-bar">
          <span class="act-title">{total_g} laudos processados no Histórico</span>
          <span class="bcnt" style="background:rgba(16,185,129,.1);color:#10B981;border:1px solid rgba(16,185,129,.3)">{norm_g} Normal</span>
          <span class="bcnt" style="background:rgba(245,158,11,.1);color:#D97706;border:1px solid rgba(245,158,11,.3)">{alt_g} Alerta</span>
          <span class="bcnt" style="background:rgba(239,68,68,.1);color:#EF4444;border:1px solid rgba(239,68,68,.3)">{alm_g} Alarme</span>
        </div>""", unsafe_allow_html=True)
    with c2:
        st.download_button("⬇️  Baixar Excel", data=excel_bytes,
                           file_name=st.session_state.get("excel", "Laudos_Historico.xlsx"),
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           use_container_width=True, type="primary")
    with c3:
        if st.button("➕ Inserir mais PDFs", use_container_width=True):
            for k in ["laudos","processado","excel"]:
                st.session_state.pop(k, None)
            st.rerun()
    with c4:
        components.html(f"""
        <style>
        * {{ margin:0; padding:0; box-sizing:border-box; }}
        body {{ background:transparent; font-family:'Barlow',sans-serif; }}
        button {{
            width:100%; height:42px; 
            background: #FFFFFF;
            color:#0055A5; border:1px solid #CBD5E1; border-radius:4px;
            font-size:13px; font-weight:700; cursor:pointer;
            display:flex; align-items:center; justify-content:center; gap:6px;
            transition:all .2s; letter-spacing:.3px;
        }}
        button:hover {{ background: #F8FAFC; border-color:#0055A5; }}
        </style>
        <button onclick="entrarTelaCheia()">⛶ &nbsp;Apresentar</button>
        <script>
        function entrarTelaCheia() {{
            try {{
                var p = window.parent;
                var doc = p.document;
                var el  = doc.documentElement;
                var fn = el.requestFullscreen || el.webkitRequestFullscreen || el.mozRequestFullScreen || el.msRequestFullscreen;
                if (fn) fn.call(el);
                function hide() {{
                    var sels = ['header[data-testid="stHeader"]','[data-testid="stToolbar"]','[data-testid="stDecoration"]','[data-testid="stStatusWidget"]','#MainMenu', 'footer'];
                    sels.forEach(function(s) {{ var elem = doc.querySelector(s); if (elem) elem.style.cssText = 'display:none!important'; }});
                    var bc = doc.querySelector('.block-container');
                    if (bc) {{ bc.style.paddingTop='0.3rem'; bc.style.paddingBottom='0'; }}
                }}
                hide(); setTimeout(hide, 500); setTimeout(hide, 1200);
            }} catch(e) {{ console.log('Fullscreen error:', e); }}
        }}
        </script>
        """, height=42)

    # ── CABEÇALHO AZUL ──
    ultima = str(df["Data de coleta"].dropna().iloc[-1]) if not df["Data de coleta"].dropna().empty else "—"
    st.markdown(f"""
    <div class="db-header">
      <div style="display:flex;align-items:center;gap:12px">
        <div class="db-logo" style="width:42px; height:42px; background:transparent; padding:0; border:none;">
            <img src="{LOGO_B64}" style="width:100%; height:100%; object-fit:contain;">
        </div>
        <div>
          <div class="db-title">MONITORAMENTO DE ANÁLISE DE ÓLEO</div>
          <div class="db-sub">Gerdau Charqueadas · Engenharia de Manutenção </div>
        </div>
     

    # ── FILTROS DA TELA ──
    setores = ["Todos"] + sorted([str(x) for x in df["Cod2"].dropna().unique().tolist()])
    anos    = ["Todos"] + sorted([str(int(x)) for x in df["Ano coleta"].dropna().unique().tolist() if pd.notna(x)], reverse=True)
    meses_disp = ["Todos"] + [MESES_F.get(int(m), "") for m in sorted(df["Mês coleta"].dropna().unique().tolist()) if pd.notna(m)]

    fc1, fc2, fc3 = st.columns([2, 1, 1])
    with fc1: setor_sel = st.selectbox("🏭  SETOR", setores, label_visibility="visible") 
    with fc2: ano_sel   = st.selectbox("📅  ANO",          anos,    label_visibility="visible")
    with fc3: mes_sel   = st.selectbox("📆  MÊS",          meses_disp, label_visibility="visible")

    dff = df.copy()
    if setor_sel != "Todos": dff = dff[dff["Cod2"] == setor_sel]
    if ano_sel   != "Todos": dff = dff[dff["Ano coleta"] == int(ano_sel)]
    if mes_sel   != "Todos":
        mn = {v: k for k, v in MESES_F.items()}[mes_sel]
        dff = dff[dff["Mês coleta"] == mn]

    total = len(dff)
    norm  = len(dff[dff["Status"]=="Normal"])
    alt   = len(dff[dff["Status"]=="Alerta"])
    alm   = len(dff[dff["Status"]=="Alarme"])
    
    pn = f"{norm/total*100:.0f}%" if total else "0%"
    pa = f"{alt/total*100:.0f}%"  if total else "0%"
    pm = f"{alm/total*100:.0f}%"  if total else "0%"

    # ── CARTÕES GRANDES DE RESUMO (KPIs) ──
    kc1, kc2, kc3 = st.columns([2, 1, 1])
    
    with kc1:
        st.markdown(f"""
        <div style="display:flex; gap:1rem;">
          <div class="kpi-tot kpi" style="flex:1;">
            <div class="kv">{total}</div>
            <div class="kl">TOTAL DE ATIVOS</div>
            <div class="ks">laudos no período filtrado</div>
          </div>
          <div class="kpi-nor kpi" style="flex:1;">
            <div class="kv">{norm}</div>
            <div class="kl">NORMAL</div>
            <div class="ks">{pn} dos laudos</div>
          </div>
        </div>""", unsafe_allow_html=True)

    with kc2:
        st.markdown(f"""
        <div class="kpi-alt kpi">
          <div class="kv">{alt}</div>
          <div class="kl">ALERTA</div>
          <div class="ks">{pa} dos laudos</div>
        </div>""", unsafe_allow_html=True)

    with kc3:
        st.markdown(f"""
        <div class="kpi-alm kpi">
          <div class="kv">{alm}</div>
          <div class="kl">ALARME</div>
          <div class="ks">{pm} dos laudos</div>
        </div>""", unsafe_allow_html=True)

    st.markdown("<div style='height: 2rem;'></div>", unsafe_allow_html=True)

    # ── OS 3 GRÁFICOS INFERIORES ──
    gc1, gc2, gc3 = st.columns([2, 1, 1])
    COR = {"Normal":"#10B981","Alerta":"#F59E0B","Alarme":"#EF4444"}
    LAYOUT = dict(plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)",
                  font=dict(color="#475569",family="Barlow"))

    with gc1:
        st.markdown('<div class="sc-top"><div class="st">Status por Setor</div></div>', unsafe_allow_html=True)
        if not dff.empty: 
            db = dff.groupby(["Cod2","Status"]).size().reset_index(name="n")
            db["tot"] = db.groupby("Cod2")["n"].transform("sum")
            db["pct"] = (db["n"] / db["tot"] * 100).round(1)
            fig = px.bar(db, x="pct", y="Cod2", color="Status", orientation="h",
                         barmode="stack", color_discrete_map=COR, text="pct",
                         category_orders={"Status":["Normal","Alerta","Alarme"]})
            fig.update_traces(texttemplate="%{text:.0f}%", textposition="inside",
                              insidetextanchor="middle", textfont_size=10, textfont_color="white", cliponaxis=False)
            fig.update_layout(**LAYOUT, height=220,
                margin=dict(l=10, r=20, t=20, b=30),
                xaxis=dict(showgrid=False,showticklabels=False,range=[0, 115]),
                yaxis=dict(showgrid=False,tickfont=dict(color="#0F172A",size=12,family="Barlow Condensed")),
                legend=dict(orientation="h",y=-0.2,x=0.5,xanchor="center",font=dict(color="#475569",size=10), bgcolor="rgba(0,0,0,0)"),
                bargap=0.3)
            st.plotly_chart(fig, use_container_width=True, config={"displayModeBar":False})

    with gc2:
        st.markdown('<div class="sc-top"><div class="st">Distribuição Global</div></div>', unsafe_allow_html=True)
        if not dff.empty and total > 0:
            dp = dff["Status"].value_counts().reset_index()
            dp.columns = ["Status","n"]
            fig2 = go.Figure(go.Pie(
                labels=dp["Status"], values=dp["n"], hole=0.55,
                marker_colors=[COR.get(s,"#94A3B8") for s in dp["Status"]],
                textinfo="percent", textfont_size=11, textfont_color="white",
                hovertemplate="%{label}: %{value}<extra></extra>"
            ))
            fig2.add_annotation(text=f"<b>{total}</b>", x=0.5, y=0.5,
                font=dict(size=24,color="#0F172A",family="Barlow Condensed"), showarrow=False)
            fig2.update_layout(**LAYOUT, height=220,
                margin=dict(l=10, r=10, t=35, b=50),
                showlegend=True,
                legend=dict(orientation="h",y=-0.3,x=0.5,xanchor="center", font=dict(color="#475569",size=10), bgcolor="rgba(0,0,0,0)"))
            st.plotly_chart(fig2, use_container_width=True, config={"displayModeBar":False})

    with gc3:
        st.markdown('<div class="sc-top"><div class="st">Coletas do Histórico (Meses)</div></div>', unsafe_allow_html=True)
        if not dff.empty:
            dm = dff.groupby("Mês coleta").size().reset_index(name="n").sort_values("Mês coleta")
            dm["label"] = dm["Mês coleta"].map(lambda x: MESES.get(int(x),"") if pd.notna(x) else "")
            max_y = dm["n"].max() if not dm.empty else 10
            
            fig3 = go.Figure(go.Bar(
                x=dm["label"], y=dm["n"], marker_color="#0055A5", 
                text=dm["n"], textposition="outside",
                textfont=dict(color="#475569",size=10),
                hovertemplate="%{x}: %{y} laudos<extra></extra>"
            ))
            fig3.update_traces(cliponaxis=False)
            fig3.update_layout(**LAYOUT, height=220,
                margin=dict(l=10, r=10, t=25, b=30),
                xaxis=dict(showgrid=False,tickfont=dict(color="#64748B",size=10)),
                yaxis=dict(showgrid=False,showticklabels=False, range=[0, max_y * 1.3]))
            st.plotly_chart(fig3, use_container_width=True, config={"displayModeBar":False})

    # ── TABELA FINAL ──
    st.markdown('<div class="sc"><div class="st">Equipamentos com Desvio no Histórico</div>', unsafe_allow_html=True)

    dtab = dff[dff["Status"].isin(["Alarme","Alerta"])].sort_values(["Ano coleta", "Mês coleta", "Status"], ascending=[False, False, True])
    if dtab.empty:
        st.markdown('<p style="color:#10B981;text-align:center;padding:16px;font-weight:600;">✅ Nenhum equipamento com desvio neste período.</p>', unsafe_allow_html=True)
    else:
        rows = ""
        for _, r in dtab.iterrows():
            st_val   = r.get("Status","")
            bcls     = "ba" if st_val=="Alarme" else "bl"
            ferro    = to_num(r.get("Ferro (ppm)"))
            agua     = to_num(r.get("Água (ppm)"))
            visc     = to_num(r.get("Visc 40 (cSt)"))
            
            fc = "td-red" if ferro and ferro>30 else ("td-yel" if ferro and ferro>15 else "")
            ac = "td-red" if agua  and agua>800  else ("td-yel" if agua  and agua>200  else "")
            vc = "td-yel" if visc else ""
            
            rows += f"""<tr>
              <td class="td-set">{r.get('Cod2','')}</td>
              <td style="color:#64748B;font-size:11px">{r.get('Cod3','')}</td>
              <td class="td-eq">{str(r.get('Equipamento Descrição 1','') or '')}</td>
              <td><span class="badge {bcls}">{st_val}</span></td>
              <td class="td-par">{str(r.get('Parâmetros em Alarme','') or '—')}</td>
              <td class="td-obs">{str(r.get('Observações','') or '')[:55]}</td>
              <td class="{fc}">{ferro if ferro is not None else '—'}</td>
              <td class="{ac}">{agua  if agua  is not None else '—'}</td>
              <td class="{vc}">{visc  if visc  is not None else '—'}</td>
              <td class="td-data">{str(r.get('Data de coleta','') or '')[:10]}</td>
            </tr>"""
            
        st.markdown(f"""
        <div style="overflow:hidden;border-radius:4px">
        <table class="tbl"><thead><tr>
          <th>Setor</th><th>Área</th><th>Equipamento</th><th>Status</th>
          <th>Parâmetro</th><th>Observações</th>
          <th>Ferro (ppm)</th><th>Água (ppm)</th><th>Visc 40</th><th>Coleta</th>
        </tr></thead><tbody>{rows}</tbody></table></div>""", unsafe_allow_html=True)

    st.markdown('</div>', unsafe_allow_html=True)

    st.markdown(f"""
    <div style="text-align:center;margin-top:10px;font-size:9px;color:#94A3B8;letter-spacing:.4px">
      SKF · Gerdau Charqueadas · Engenharia de Manutenção &nbsp;·&nbsp;
      Desenvolvido por Douglas Brum &nbsp;·&nbsp; {datetime.now().strftime('%d/%m/%Y')}
    </div>""", unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════════════
# ROTEADOR PRINCIPAL (A porta de Entrada)
# ══════════════════════════════════════════════════════════════════════════════
if "processado" not in st.session_state:
    if "tentativa_carga" not in st.session_state:
        st.session_state["tentativa_carga"] = True
        historico = carregar_historico()
        if historico:
            st.session_state["laudos"] = historico
            st.session_state["processado"] = True
            st.rerun()

if st.session_state.get("processado"):
    render_dashboard(st.session_state["laudos"])
else:
    render_upload()
